import dns.resolver
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, colors, Color, PatternFill
import getpass
from datetime import datetime
import pyodbc
import ipaddress

#You will understoot nothin from this whole code, because you have no idea about DB strcuture and task which I try to solve )))
#I'm trying here to analyse data from SQL, group it and put result to excel file


# START
start_time = datetime.now()
dbUser=input("Please Enter username: ")
dbPass= getpass.getpass('Password:')


currDir="c://Users/rpomazanov//Documents//VS_1//Projects//FrankfurtNetwork_NEW//Streams//App migration//Analysis//VMs_Apps//"

cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER=10.76.128.76,10400;DATABASE=InventoryFeed;UID=' +dbUser+';PWD='+dbPass)
cursor = cnxn.cursor()

#Collect all apps
dctApps=dict()
cursor.execute("SELECT Name, ITOwnerGroup, ITOwner, BusinessOwner, BusinessLine, TypeOfApplication FROM dbo.Export_Applications")

# Values meaning:
# 1 - ITOwnerGroup
# 2 - ITOwner
# 3 - BusinessOwner
# 4 - BusinessLine
# 5 - Type of Application

for row in cursor:
    lstTmp=[row[1], row[2], row [3], row[4], row[5]]
    dctApps[row[0]]=lstTmp


# Get Interfaces
dctSubnets=dict()
wb=load_workbook(currDir+"SophosSubnets.xlsx")
ws=wb.active
for i in range(2,ws.max_row+1):
    #Dictionary - Subnet, Interface, Interfaces description, Interface comment
    dctSubnets[ws["B"+str(i)].value]=[ws["A"+str(i)].value,ws["C"+str(i)].value,ws["D"+str(i)].value]

#Collect all FRA servers
dctServers=dict()
cursor.execute("SELECT HostName, Name,IPs,Description,ApplicationsHostedOnThisServer,  OwnerGroup, PrimaryOwner, ESXCLusterName,    ESXHost, SecondaryOwner FROM dbo.Export_Servers WHERE Location LIKE 'Frankfurt%' AND PrimaryRole <> 'Retired' AND Zone = 'Internal' AND PrimaryRole <> 'Template' AND HostName<>'photon-machine' ")
sSubnet=""
for row in cursor: 
    sDNS=row[0]
    sName=row[1]
    sIP=row[2]
    sDscr=row[3]
    sApp=row[4]
    sOG=row[5]
    sPO=row[6]
    sESXCL=row[7]
    sESXHost=row[8]
    sSO=row[9]
    #Some servers do not have IP address in infra
    #if has
    if isinstance(sIP,str):
        IPsrc="Infra"
    else:  
        #IF not - Try to resolve via DNS  
        try:
            dns_result=dns.resolver.query(row[0])
            sIP=str(dns_result[0])
            IPsrc="DNS"
        # If no success with DNS resolving
        except:
            sIP="N\A"  
            IPsrc="No_IP"     
    if len(sIP.split(","))>1:
        sIP=sIP.split(",")[0]
    # if IP has been found, let's try to idenrify the subnet by seeking on subnet dictionary if IP overlap or not
    if sIP!="N\A":
        for s in dctSubnets:
            net=s
            net_ip = ipaddress.IPv4Network(net)
            ip_add = ipaddress.IPv4Network(sIP)
            if net_ip.overlaps(ip_add):
                sSubnet=s
                sInt=dctSubnets[s][0]
                sIntDes=dctSubnets[s][1]
                sIntComment=dctSubnets[s][2]
    else:
        sSubnet="No Subnet info"
        sInt="N\A"
        sIntDes="N\A"
        sIntComment="N\A"

    dctServers[sName]=[sDNS, sIP, sSubnet, sDscr, sOG, sPO, sESXCL, sESXHost,IPsrc,sApp,sInt,sIntDes,sIntComment, sSO]
#################################################    
#################Servers by Subnet###############
#################################################

wb=Workbook()
wb.create_sheet("ServersBySubnets")
ws=wb["ServersBySubnets"]
row=1
# Seeking for servers wich are located in each subnet
for i in dctSubnets:
    ws["A"+str(row)]="Interface"
    ws["A"+str(row)].font=Font(bold=True,color=colors.RED)
    ws["A"+str(row)].fill=PatternFill("solid",fgColor=colors.YELLOW)
    ws["B"+str(row)]=dctSubnets[i][0]
    ws["B"+str(row)].font=Font(color=colors.RED)
    ws["B"+str(row)].fill=PatternFill("solid",fgColor=colors.YELLOW)
    ws["C"+str(row)]="Number of servers based on infra"
    ws["C"+str(row)].font=Font(bold=True,color=colors.GREEN)
    firstrow=row    
    row=row+1
    ws["A"+str(row)]="Subnet"
    ws["A"+str(row)].font=Font(bold=True,color=colors.RED)
    ws["A"+str(row)].fill=PatternFill("solid",fgColor=colors.YELLOW)

    ws["B"+str(row)]=str(i)
    ws["B"+str(row)].font=Font(color=colors.RED)
    ws["B"+str(row)].fill=PatternFill("solid",fgColor=colors.YELLOW)
    
    row=row+1
    ws["A"+str(row)]="Description"
    ws["A"+str(row)].font=Font(bold=True,color=colors.RED)
    ws["B"+str(row)]=dctSubnets[i][1]
    ws["B"+str(row)].font=Font(color=colors.RED)
    
    row=row+1
    ws["A"+str(row)]="Comment"
    ws["A"+str(row)].font=Font(bold=True,color=colors.RED)
    ws["B"+str(row)]=dctSubnets[i][2]
    ws["B"+str(row)].font=Font(color=colors.RED)
    
    row=row+1
    ws["A"+str(row)]="Server Name"
    ws["B"+str(row)]="Server IP"
    ws["C"+str(row)]="Primary Owner"
    ws["D"+str(row)]="Application on Server"
    ws["E"+str(row)]="Application Businness owner"
    ws["F"+str(row)]="Application IT owner"
    ws["G"+str(row)]="Description"
    ws["A"+str(row)].font=Font(bold=True)
    ws["B"+str(row)].font=Font(bold=True)
    ws["C"+str(row)].font=Font(bold=True)
    ws["D"+str(row)].font=Font(bold=True)
    ws["E"+str(row)].font=Font(bold=True)
    ws["F"+str(row)].font=Font(bold=True)    
    ws["H"+str(row)]="Secondary Owner"
    row=row+1
    #Looking into servers dictioanray and if current subnet (i) the same as for server - fill data....
    for j in dctServers:
        if i==dctServers[j][2]:
            ws["A"+str(row)]=j                 # Server name 
            ws["B"+str(row)]=dctServers[j][1]  # Server IP 
            ws["C"+str(row)]=dctServers[j][5]  # Primary Owner
            ws["H"+str(row)]=dctServers[j][13]  # Secondary Owner
            ws["D"+str(row)]=dctServers[j][9]  # Application 
            ws["G"+str(row)]=dctServers[j][3]  # Application 
            try:
                # and try to find some information about application
                for k in dctApps:
                    if  k==dctServers[j][9]:
                        ws["E"+str(row)]=dctApps[k][2]     # Application Businness owner
                        ws["F"+str(row)]=dctApps[k][1]     # Application IT owner
            except:
                #Sometimes there is no any info about application
                    ws["E"+str(row)]="NoAppInfo"
                    ws["F"+str(row)]="NoAppInfo"
    

            row=row+1
    ws["D"+str(firstrow)]=row-firstrow-5
    ws["D"+str(firstrow)].font=Font(bold=True,color=colors.GREEN)
    row=row+1
#################"AUTOFITT"CODE ########################################
for col in range(ord("A"),ord("G")):
    col_max_width = 0
    for line in range(1,ws.max_row+1):
        if len( str(ws[chr(col)+str(line)].value) ) > col_max_width:
            col_max_width = len(str(ws[chr(col)+str(line)].value))
    ws.column_dimensions[chr(col)].width = col_max_width


#################################################    
#######Applications by Servers Subnet############
#################################################


### Create new dct with only FRA apps
dctFraApps=dict()

# Check if  server attached to specific application. If so - add to new dct
for i in dctServers:
    try:
        dctFraApps[dctServers[i][9]]=dctApps[dctServers[i][9]]
    except:
        print ("There is no info about " +dctServers[i][9]  )


wb.create_sheet("AppsOnSubnets")
ws=wb["AppsOnSubnets"]
row=1
for i in dctFraApps:


        ws["A"+str(row)]="Application name"
        ws["A"+str(row)].font=Font(bold=True,color=colors.RED)
        ws["B"+str(row)]=str(i)
        ws["B"+str(row)].font=Font(color=colors.RED)
        ws["A"+str(row)].fill=PatternFill("solid",fgColor=colors.YELLOW)
        ws["B"+str(row)].fill=PatternFill("solid",fgColor=colors.YELLOW)
        
        row=row+1
        ws["A"+str(row)]="App IT Owner"
        ws["A"+str(row)].font=Font(bold=True,color=colors.RED)
        ws["B"+str(row)]=dctFraApps[i][1]
        ws["B"+str(row)].font=Font(color=colors.RED)
        
        row=row+1
        ws["A"+str(row)]="App Business Owner"
        ws["A"+str(row)].font=Font(bold=True,color=colors.RED)
        ws["B"+str(row)]=dctFraApps[i][2]
        ws["B"+str(row)].font=Font(color=colors.RED)
        
        row=row+1
        ws["A"+str(row)]="App Business Line"
        ws["A"+str(row)].font=Font(bold=True,color=colors.RED)
        ws["B"+str(row)]=dctFraApps[i][3]
        ws["B"+str(row)].font=Font(color=colors.RED)
        row=row+1

        ws["A"+str(row)]="Server Name"
        ws["B"+str(row)]="Server IP"
        ws["C"+str(row)]="Primary Owner"
        ws["D"+str(row)]="Server subnet"
        ws["E"+str(row)]="Sophos Interface"
        ws["F"+str(row)]="Interface Description"
        ws["G"+str(row)]="Interface Comment"
        ws["A"+str(row)].font=Font(bold=True)
        ws["B"+str(row)].font=Font(bold=True)
        ws["C"+str(row)].font=Font(bold=True)
        ws["D"+str(row)].font=Font(bold=True)
        ws["E"+str(row)].font=Font(bold=True)
        ws["F"+str(row)].font=Font(bold=True)
        ws["G"+str(row)].font=Font(bold=True)
        row=row+1

        for j in  dctServers:
            if i==dctServers[j][9]:
                ws["A"+str(row)]=j     # SErver name
                ws["B"+str(row)]=dctServers[j][1]  # Server IP  
                ws["C"+str(row)]=dctServers[j][5]  # Primary Owner
                ws["D"+str(row)]=dctServers[j][2]  # Subnet
                ws["E"+str(row)]=dctServers[j][10] # Sophos Interface 
                ws["F"+str(row)]=dctServers[j][11] # Sophos Interface description
                ws["G"+str(row)]=dctServers[j][12] # Sophos Interface comment
                row=row+1
        row=row+1
#################"AUTOFITT"CODE ########################################
for col in range(ord("A"),ord("H")):
    col_max_width = 0
    for line in range(1,ws.max_row+1):
        if len( str(ws[chr(col)+str(line)].value) ) > col_max_width:
            col_max_width = len(str(ws[chr(col)+str(line)].value))
    ws.column_dimensions[chr(col)].width = col_max_width



wb.save(currDir+"ServersBySubnets"+str(datetime.now().date())+str(datetime.now().time()).replace(":",'')+".xlsx")


end_time=datetime.now() 
print ("Script execution time: " + str(end_time-start_time))




######################################################################
#Code for getting info from Subnet on Sophos and interfaces on Sophos#
######################################################################

#Create all subnets from Sophos
# dctSubnets=dict()
# wb=load_workbook(currDir+"RAW data.xlsx")
# ws=wb.active
# for i in range(2,ws.max_row):
#     ifc=ws["A"+str(i)].value
#     ipadd=ws["B"+str(i)].value
#     ipadd=ipadd[-(len(ipadd)-5):]
#     mask=ws["D"+str(i)].value
#     mask=mask[-(len(mask)-5):]
#     dctSubnets[ifc]=ipaddress.IPv4Interface("{}/{}".format(ipadd,mask)).network


# ##Interfaces with masks
# dctIntSophos=dict()
# wb=load_workbook(currDir+"SOPHOS interfaces v6.xlsx")
# ws=wb.active
# for i in range (2, ws.max_row):
#     ifc=ws["A"+str(i)].value
#     dsc=ws["D"+str(i)].value
#     cmnt=ws["E"+str(i)].value
#     dctIntSophos[ifc]=[dctSubnets[ifc],dsc,cmnt]


# wwb=Workbook()
# wwb.create_sheet("Interfaces")
# wws=wwb["Interfaces"]
# row=1
# for i in dctIntSophos:
#     row=row+1
#     wws["A"+str(row)]=i
#     wws["B"+str(row)]=str(dctIntSophos[i][0])
#     wws["C"+str(row)]=dctIntSophos[i][1]
#     wws["D"+str(row)]=dctIntSophos[i][2]

#wwb.save(currDir+"SophosSubnets.xlsx")    
